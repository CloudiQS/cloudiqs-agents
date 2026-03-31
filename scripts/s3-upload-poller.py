#!/usr/bin/env python3
"""
Smart S3 Upload Poller for CloudiQS Engine.

Watches the S3 uploads bucket for new files.
Auto-detects columns: company, email, website, contact, campaign, etc.
Posts each row to bridge /ingest endpoint.
Handles CSV, XLSX, TXT, any column order, missing headers.

Design: your team drops a file with whatever data they have.
The poller figures out what each column is. No hardcoded format.
"""

import os
import re
import csv
import json
import logging
import subprocess
from pathlib import Path
from typing import Dict, List, Optional

try:
    import requests
except ImportError:
    import urllib.request
    # Fallback for systems without requests
    class requests:
        @staticmethod
        def post(url, json=None):
            data = __import__('json').dumps(json).encode()
            req = urllib.request.Request(url, data=data, headers={"Content-Type": "application/json"})
            resp = urllib.request.urlopen(req)
            class R:
                status_code = resp.status
                text = resp.read().decode()
            return R()

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(message)s")
logger = logging.getLogger("s3-poller")

# All names from environment or derive from stack name
# deploy.sh sets these. Fallback to defaults for manual runs.
_STACK = os.environ.get("STACK_NAME", "cloudiqs-engine")
_ACCOUNT = os.environ.get("AWS_ACCOUNT", "")
if not _ACCOUNT:
    try:
        result = subprocess.run(
            ["aws", "sts", "get-caller-identity", "--query", "Account", "--output", "text"],
            capture_output=True, text=True, timeout=10,
        )
        _ACCOUNT = result.stdout.strip()
    except Exception:
        _ACCOUNT = "unknown"

BUCKET = os.environ.get("S3_BUCKET", f"{_STACK}-uploads-{_ACCOUNT}")
REGION = os.environ.get("AWS_REGION", "eu-west-1")
BRIDGE_URL = os.environ.get("BRIDGE_URL", "http://localhost:8787/ingest")
PROCESSED_DIR = os.path.expanduser("~/.openclaw/processed-uploads")
TEMP_DIR = "/tmp/s3-poller"

# Column detection patterns
PATTERNS = {
    "email": re.compile(r"^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$"),
    "website": re.compile(r"(https?://|www\.|\.(com|co\.uk|org|io|net|ai))"),
    "linkedin": re.compile(r"linkedin\.com", re.I),
    "postcode": re.compile(r"^[A-Z]{1,2}\d[A-Z\d]?\s*\d[A-Z]{2}$", re.I),
    "phone": re.compile(r"^[\+\d\s\(\)-]{8,}$"),
    "companies_house": re.compile(r"^\d{7,8}$"),
}

# Header name mapping (fuzzy match)
HEADER_MAP = {
    "company": ["company", "company name", "organisation", "organization", "business", "name", "account"],
    "email": ["email", "e-mail", "mail", "email address", "contact email"],
    "website": ["website", "url", "web", "domain", "site"],
    "contact": ["contact", "contact name", "full name", "person", "decision maker", "dm"],
    "job_title": ["title", "job title", "role", "position", "job"],
    "campaign": ["campaign", "vertical", "segment", "type", "category"],
    "phone": ["phone", "telephone", "tel", "mobile"],
    "linkedin": ["linkedin", "linkedin url", "li url", "li"],
    "notes": ["notes", "comments", "description", "info"],
}


def detect_column_type(header: str, values: List[str]) -> str:
    """Detect what type of data a column contains."""
    header_lower = header.lower().strip()

    # First try header name matching
    for col_type, aliases in HEADER_MAP.items():
        if header_lower in aliases:
            return col_type

    # Then try data pattern matching (sample first 5 non-empty values)
    samples = [v.strip() for v in values if v and v.strip()][:5]
    if not samples:
        return "unknown"

    # Count pattern matches
    for col_type, pattern in PATTERNS.items():
        matches = sum(1 for s in samples if pattern.search(s))
        if matches >= len(samples) * 0.6:  # 60% match threshold
            return col_type

    # If values are short capitalised words (2-3 words), likely company or contact
    avg_words = sum(len(s.split()) for s in samples) / len(samples)
    if avg_words <= 4 and all(s[0].isupper() for s in samples if s):
        # First unmatched column of this type is company, second is contact
        return "company_or_contact"

    return "unknown"


def parse_file(filepath: str) -> List[Dict]:
    """Parse any file format into a list of row dicts with detected column types."""
    ext = Path(filepath).suffix.lower()

    if ext == ".xlsx":
        return parse_xlsx(filepath)
    elif ext in (".csv", ".tsv"):
        return parse_csv(filepath)
    else:
        return parse_text(filepath)


def parse_csv(filepath: str) -> List[Dict]:
    """Parse CSV with auto-detection."""
    rows = []
    with open(filepath, newline="", encoding="utf-8-sig") as f:
        # Sniff dialect
        sample = f.read(4096)
        f.seek(0)

        try:
            dialect = csv.Sniffer().sniff(sample)
            has_header = csv.Sniffer().has_header(sample)
        except csv.Error:
            dialect = csv.excel
            has_header = True

        reader = csv.reader(f, dialect)
        all_rows = list(reader)

        if not all_rows:
            return []

        if has_header:
            headers = [h.strip() for h in all_rows[0]]
            data_rows = all_rows[1:]
        else:
            # No header - generate column names
            headers = [f"col_{i}" for i in range(len(all_rows[0]))]
            data_rows = all_rows

        # Detect column types
        col_types = {}
        for i, header in enumerate(headers):
            values = [row[i] if i < len(row) else "" for row in data_rows]
            col_types[i] = detect_column_type(header, values)

        # Resolve company_or_contact: first is company, second is contact
        seen_company = False
        for i in sorted(col_types.keys()):
            if col_types[i] == "company_or_contact":
                if not seen_company:
                    col_types[i] = "company"
                    seen_company = True
                else:
                    col_types[i] = "contact"

        # Build row dicts
        for row in data_rows:
            if not any(cell.strip() for cell in row):
                continue  # Skip empty rows

            record = {}
            for i, val in enumerate(row):
                if i in col_types and col_types[i] != "unknown":
                    record[col_types[i]] = val.strip()

            if record.get("company"):
                rows.append(record)

    return rows


def parse_xlsx(filepath: str) -> List[Dict]:
    """Parse Excel file."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        logger.error("openpyxl not installed - cannot parse XLSX")
        return []

    wb = load_workbook(filepath, read_only=True)
    ws = wb.active

    all_rows = []
    for row in ws.iter_rows(values_only=True):
        all_rows.append([str(cell) if cell is not None else "" for cell in row])

    if not all_rows:
        return []

    # Treat first row as header, detect types same as CSV
    headers = [h.strip() for h in all_rows[0]]
    data_rows = all_rows[1:]

    col_types = {}
    for i, header in enumerate(headers):
        values = [row[i] if i < len(row) else "" for row in data_rows]
        col_types[i] = detect_column_type(header, values)

    seen_company = False
    for i in sorted(col_types.keys()):
        if col_types[i] == "company_or_contact":
            if not seen_company:
                col_types[i] = "company"
                seen_company = True
            else:
                col_types[i] = "contact"

    rows = []
    for row in data_rows:
        if not any(cell.strip() for cell in row if cell):
            continue
        record = {}
        for i, val in enumerate(row):
            if i in col_types and col_types[i] != "unknown":
                record[col_types[i]] = val.strip()
        if record.get("company"):
            rows.append(record)

    return rows


def parse_text(filepath: str) -> List[Dict]:
    """Parse plain text file - one company per line."""
    rows = []
    with open(filepath, encoding="utf-8-sig") as f:
        for line in f:
            company = line.strip()
            if company and not company.startswith("#"):
                rows.append({"company": company})
    return rows


def submit_to_bridge(record: Dict) -> bool:
    """POST a record to the bridge /ingest endpoint."""
    payload = {
        "company": record.get("company", ""),
        "campaign": record.get("campaign", "triage"),
        "website": record.get("website", ""),
        "contact": record.get("contact", ""),
        "email": record.get("email", ""),
        "job_title": record.get("job_title", ""),
        "notes": record.get("notes", ""),
        "source": "s3-upload",
    }

    try:
        r = requests.post(BRIDGE_URL, json=payload)
        if r.status_code in (200, 201):
            return True
        logger.error(f"Bridge rejected {payload['company']}: {r.text[:100]}")
    except Exception as e:
        logger.error(f"Bridge error for {payload['company']}: {e}")
    return False


def main():
    os.makedirs(PROCESSED_DIR, exist_ok=True)
    os.makedirs(TEMP_DIR, exist_ok=True)

    # List files in S3 (exclude cvs/ subfolder - that is handled by cv-poller)
    try:
        result = subprocess.run(
            ["aws", "s3", "ls", f"s3://{BUCKET}/", "--region", REGION],
            capture_output=True, text=True, timeout=30,
        )
    except Exception as e:
        logger.error(f"S3 list failed: {e}")
        return

    if result.returncode != 0:
        logger.error(f"S3 list error: {result.stderr[:200]}")
        return

    for line in result.stdout.strip().split("\n"):
        if not line.strip():
            continue

        parts = line.split()
        if len(parts) < 4:
            continue

        filename = parts[3]

        # Skip directories and processed files
        if filename.endswith("/"):
            continue
        if os.path.exists(os.path.join(PROCESSED_DIR, filename)):
            continue

        logger.info(f"Processing: {filename}")

        # Download
        local_path = os.path.join(TEMP_DIR, filename)
        dl = subprocess.run(
            ["aws", "s3", "cp", f"s3://{BUCKET}/{filename}", local_path, "--region", REGION],
            capture_output=True, text=True, timeout=60,
        )
        if dl.returncode != 0:
            logger.error(f"Download failed: {filename}")
            continue

        # Parse
        records = parse_file(local_path)
        logger.info(f"Parsed {len(records)} records from {filename}")

        # Submit each record
        success = 0
        for record in records:
            if submit_to_bridge(record):
                success += 1

        logger.info(f"Submitted {success}/{len(records)} from {filename}")

        # Only mark as processed if most rows succeeded
        # This ensures files get retried if bridge was down
        if len(records) == 0:
            # Empty file, mark as done
            Path(os.path.join(PROCESSED_DIR, filename)).touch()
        elif success >= len(records) * 0.8:
            # 80%+ succeeded, mark as done
            Path(os.path.join(PROCESSED_DIR, filename)).touch()
            logger.info(f"Marked as processed: {filename}")
        else:
            # Too many failures, will retry next run
            logger.warning(
                f"Only {success}/{len(records)} submitted from {filename}. "
                f"NOT marking as processed. Will retry next run."
            )

        # Cleanup temp
        try:
            os.remove(local_path)
        except OSError:
            pass


if __name__ == "__main__":
    main()
